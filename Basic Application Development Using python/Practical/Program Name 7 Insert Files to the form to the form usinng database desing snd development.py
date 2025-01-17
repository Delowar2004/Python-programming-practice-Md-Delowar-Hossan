from openpyxl import *
from tkinter import *

wb = Workbook()
sheet = wb.create_sheet("Sheet1")

def excel():
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 20
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 40
    sheet.column_dimensions['G'].width = 50

    sheet.cell(row=1, column=1).value = "Name"
    sheet.cell(row=1, column=2).value = "Course"
    sheet.cell(row=1, column=3).value = "Semester"
    sheet.cell(row=1, column=4).value = "Form Number"
    sheet.cell(row=1, column=5).value = "Contact Number"
    sheet.cell(row=1, column=6).value = "Email id"
    sheet.cell(row=1, column=7).value = "Address"

def clear():
    name_field.delete(0, END)
    course_field.delete(0, END)
    sem_field.delete(0, END)
    form_no_field.delete(0, END)
    contact_no_field.delete(0, END)
    email_id_field.delete(0, END)
    address_field.delete(0, END)

def insert():
    if (name_field.get() == "" and
        course_field.get() == "" and
        sem_field.get() == "" and
        form_no_field.get() == "" and
        contact_no_field.get() == "" and
        email_id_field.get() == "" and
        address_field.get() == ""):
        print("empty input")
    else:
        current_row = sheet.max_row + 1

        sheet.cell(row=current_row, column=1).value = name_field.get()
        sheet.cell(row=current_row, column=2).value = course_field.get()
        sheet.cell(row=current_row, column=3).value = sem_field.get()
        sheet.cell(row=current_row, column=4).value = form_no_field.get()
        sheet.cell(row=current_row, column=5).value = contact_no_field.get()
        sheet.cell(row=current_row, column=6).value = email_id_field.get()
        sheet.cell(row=current_row, column=7).value = address_field.get()

        wb.save(r'2:\excel.xlsx')
        name_field.focus_set()
        clear()

root = Tk()
root.configure(background='light green')
root.title("Registration Form")
root.geometry("500x300")

excel()

heading = Label(root, text="Form", bg="light green")
heading.grid(row=0, column=1)

name = Label(root, text="Name", bg="light green")
name.grid(row=1, column=0)

course = Label(root, text="Course", bg="light green")
course.grid(row=2, column=0)

sem = Label(root, text="Semester", bg="light green")
sem.grid(row=3, column=0)

form_no = Label(root, text="Form No.", bg="light green")
form_no.grid(row=4, column=0)

contact_no = Label(root, text="Contact No.", bg="light green")
contact_no.grid(row=5, column=0)

email_id = Label(root, text="Email id", bg="light green")
email_id.grid(row=6, column=0)

address = Label(root, text="Address", bg="light green")
address.grid(row=7, column=0)

name_field = Entry(root)
course_field = Entry(root)
sem_field = Entry(root)
form_no_field = Entry(root)
contact_no_field = Entry(root)
email_id_field = Entry(root)
address_field = Entry(root)

name_field.grid(row=1, column=1, ipadx="100")
course_field.grid(row=2, column=1, ipadx="100")
sem_field.grid(row=3, column=1, ipadx="100")
form_no_field.grid(row=4, column=1, ipadx="100")
contact_no_field.grid(row=5, column=1, ipadx="100")
email_id_field.grid(row=6, column=1, ipadx="100")
address_field.grid(row=7, column=1, ipadx="100")

submit = Button
